#!/usr/bin/env python3
"""
Excel Timestamp Synchronizer - Single program with JSON configuration
"""

import json
import pandas as pd
import numpy as np
from pathlib import Path
import logging
import sys
import argparse
from datetime import datetime, date, time
from openpyxl.utils import get_column_letter


class ExcelTimestampSynchronizer:
    """
    Synchronizes data from multiple Excel files based on timestamp matching.
    """

    def __init__(self, reference_file, reference_timestamp_column, tolerance_seconds=0,
                 duplicate_strategy='mean', reference_headers=None):
        """Initialize the synchronizer.

        Args:
            reference_file (str): Path to the reference Excel file
            reference_timestamp_column (str): Name of the timestamp column in reference file (or Excel letter)
            tolerance_seconds (int): Tolerance in seconds for timestamp matching (default: 0 for exact match)
            duplicate_strategy (str): How to handle duplicate timestamps ('mean', 'first', 'last', 'max', 'min')
            reference_headers (dict|None): Optional mapping to rename headers in the reference file
        """
        self.reference_file = reference_file
        self.reference_timestamp_column = reference_timestamp_column
        self.tolerance_seconds = tolerance_seconds
        self.duplicate_strategy = duplicate_strategy
        self.reference_data = None
        self.synchronized_data = None
        self.duplicate_warnings = []
        self.reference_headers = reference_headers or {}

        # Setup logging
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        self.logger = logging.getLogger(__name__)

    # -----------------------------
    # Helpers for column selection
    # -----------------------------
    def _apply_headers(self, df: pd.DataFrame, headers_mapping):
        """Optionally rename columns using a mapping provided in config.
        Mapping keys can be:
        - Excel letters (e.g., 'A', 'B', 'AA') -> rename column at that position
        - 0-based indices as int or numeric string (e.g., 0, "0")
        - Existing column names -> directly rename
        """
        if not headers_mapping:
            return df

        new_columns = list(df.columns)

        for key, new_name in headers_mapping.items():
            if not isinstance(new_name, str) or new_name.strip() == "":
                continue

            try:
                # By existing name
                if isinstance(key, str) and key in df.columns:
                    idx = list(df.columns).index(key)
                # By Excel letter
                elif isinstance(key, str) and self._is_excel_col_letter(key):
                    idx = self._excel_col_to_index(key)
                # By numeric index (int or numeric string)
                elif isinstance(key, int) or (isinstance(key, str) and key.isdigit()):
                    idx = int(key)
                else:
                    self.logger.warning("Unsupported header mapping key: %s", key)
                    continue

                if 0 <= idx < len(new_columns):
                    new_columns[idx] = new_name
                else:
                    self.logger.warning("Header mapping index out of range: %s", key)
            except (ValueError, TypeError, IndexError) as e:  # guard around mapping parsing
                self.logger.warning("Failed applying header for key %s: %s", key, e)

        df.columns = new_columns
        return df
    def _parse_col_spec(self, spec):
        """
        Parse a column spec which can be:
        - string: column name or Excel letter, implies skip_rows=0
        - dict: { "column"|"name"|"letter": str, "skip_rows": int }
        Returns (col_token: str, skip_rows: int)
        """
        if isinstance(spec, str):
            return spec, 0
        if isinstance(spec, dict):
            col_token = spec.get('column') or spec.get('name') or spec.get('letter')
            if not isinstance(col_token, str):
                raise ValueError("Invalid column spec: missing 'column'/'name'/'letter' string")
            skip = spec.get('skip_rows', 0)
            if not isinstance(skip, int) or skip < 0:
                raise ValueError("'skip_rows' must be a non-negative integer")
            return col_token, skip
        raise ValueError("Column spec must be a string or an object with 'column' and optional 'skip_rows'")
    def _is_excel_col_letter(self, col_spec):
        """Return True if col_spec looks like an Excel column letter (A..XFD)."""
        return isinstance(col_spec, str) and col_spec.isalpha() and 1 <= len(col_spec) <= 3

    def _excel_col_to_index(self, col_spec):
        """Convert Excel column letters (A, B, ..., Z, AA, AB, ...) to 0-based index."""
        s = col_spec.upper()
        idx = 0
        for ch in s:
            if not ('A' <= ch <= 'Z'):
                raise ValueError(f"Invalid Excel column letter: {col_spec}")
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx - 1  # 0-based

    def _get_series_by_spec(self, df, col_spec):
        """
        Get a pandas Series from df by either column name or Excel letter.
        - If col_spec matches an existing column name, return df[col_spec].
        - Else if col_spec is Excel letters, return df.iloc[:, index].
        Raises KeyError if not found.
        """
        if isinstance(col_spec, str) and col_spec in df.columns:
            return df[col_spec]
        if self._is_excel_col_letter(col_spec):
            idx = self._excel_col_to_index(col_spec)
            if 0 <= idx < df.shape[1]:
                return df.iloc[:, idx]
        raise KeyError(f"Column not found by name or letter: {col_spec}")

    def _to_datetime_flexible(self, s: pd.Series) -> pd.Series:
        """Convert a Series to pandas datetime, handling time-only values gracefully.
        - First try pd.to_datetime directly.
        - If it fails and values are datetime.time, combine with a base date.
        - Else coerce values to datetime when possible.
        """
        try:
            return pd.to_datetime(s)
        except (TypeError, ValueError):
            pass

        # If series is primarily datetime.time objects
        try:
            if s.map(lambda v: (v is None) or pd.isna(v) or isinstance(v, time)).all():
                base = date(1970, 1, 1)
                return s.map(lambda v: pd.NaT if (v is None or pd.isna(v)) else datetime.combine(base, v))
        except (TypeError, ValueError):
            pass

        # Fallback: try coercion and Excel origin for numerics
        try:
            return pd.to_datetime(s, errors='coerce')
        except (TypeError, ValueError):
            # last resort: map each value
            def conv(v):
                if v is None or (isinstance(v, float) and np.isnan(v)):
                    return pd.NaT
                if isinstance(v, pd.Timestamp):
                    return v
                if isinstance(v, datetime):
                    return v
                if isinstance(v, time):
                    return datetime.combine(date(1970, 1, 1), v)
                try:
                    return pd.to_datetime(v, errors='coerce')
                except (TypeError, ValueError):
                    return pd.NaT
            return s.map(conv)

    def load_reference_data(self):
        """Load and prepare the reference Excel file."""
        try:
            self.reference_data = pd.read_excel(self.reference_file)

            # Apply optional reference headers mapping
            if isinstance(self.reference_headers, dict) and self.reference_headers:
                self.reference_data = self._apply_headers(self.reference_data, self.reference_headers)

            # Support dict for reference_timestamp_column (with optional skip_rows)
            ref_col_token = self.reference_timestamp_column
            ref_skip = 0
            if isinstance(self.reference_timestamp_column, dict):
                ref_col_token, ref_skip = self._parse_col_spec(self.reference_timestamp_column)
            
            # Apply skip_rows (drop first N rows) if any
            if ref_skip > 0:
                self.reference_data = self.reference_data.iloc[ref_skip:, :].reset_index(drop=True)

            # Ensure the reference timestamp column exists; support Excel letters (A, B, C, ...)
            if (isinstance(ref_col_token, str) and ref_col_token in self.reference_data.columns) is False:
                # Try to resolve via Excel column letters
                if self._is_excel_col_letter(ref_col_token):
                    try:
                        ser = self._get_series_by_spec(self.reference_data, ref_col_token)
                        # Create a column with the configured name so downstream code works
                        # Choose normalized name: if user provided a dict, use the token; else keep original name string
                        # Prefer the series' existing name (possibly set via headers mapping)
                        target_name = ser.name if getattr(ser, 'name', None) else (
                            ref_col_token if isinstance(self.reference_timestamp_column, dict) else self.reference_timestamp_column
                        )
                        self.reference_data[target_name] = ser
                        # Make sure self.reference_timestamp_column points to a string column name used below
                        self.reference_timestamp_column = target_name
                    except (KeyError, ValueError, IndexError) as e:
                        self.logger.error("Cannot resolve reference timestamp column '%s': %s", ref_col_token, e)
                        return False
                else:
                    self.logger.error("Reference timestamp column '%s' not found in reference file", ref_col_token)
                    return False
            else:
                # If column exists by name and was provided as dict, align attribute to the name
                if isinstance(self.reference_timestamp_column, dict):
                    self.reference_timestamp_column = ref_col_token

            # Convert timestamp column to datetime (supports time-only)
            self.reference_data[self.reference_timestamp_column] = self._to_datetime_flexible(
                self.reference_data[self.reference_timestamp_column]
            )

            self.logger.info("Loaded reference file: %s", self.reference_file)
            self.logger.info("Reference data shape: %s", self.reference_data.shape)

            return True

        except (FileNotFoundError, ValueError, KeyError, OSError) as e:
            self.logger.error("Error loading reference file: %s", e)
            return False

    def synchronize_files(self, file_configs, output_file):
        """
        Synchronize multiple Excel files with the reference timestamps.

        Args:
            file_configs (list): List of dictionaries with file configuration
            output_file (str): Path for the output synchronized Excel file
        """
        if self.reference_data is None:
            if not self.load_reference_data():
                return False

        # Start with only the reference timestamp column
        result_data = pd.DataFrame({
            self.reference_timestamp_column: self.reference_data[self.reference_timestamp_column]
        })

        for config in file_configs:
            try:
                self._process_file(config, result_data)
            except (KeyError, ValueError, OSError) as e:
                self.logger.error("Error processing file %s: %s", config['file_path'], e)
                continue

        # Save synchronized data with reference time column formatted as HH:MM:SS
        try:
            sheet_name = 'Sheet1'
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Write data
                result_data.to_excel(writer, index=False, sheet_name=sheet_name)

                # Apply time-only number format to the reference timestamp column
                ws = writer.sheets[sheet_name]
                ref_col_idx = result_data.columns.get_loc(self.reference_timestamp_column) + 1  # 1-based
                ref_col_letter = get_column_letter(ref_col_idx)

                # Skip header row (start at row 2)
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"{ref_col_letter}{row}"]
                    # Only set format if there's a value
                    if cell.value is not None:
                        cell.number_format = 'hh:mm:ss'

            self.logger.info("Synchronized data saved to: %s", output_file)
            self.synchronized_data = result_data
            return True

        except (OSError, ValueError) as e:
            self.logger.error("Error saving output file: %s", e)
            return False

    def _process_file(self, config, result_data):
        """Process a single Excel file and merge with result data."""
        file_path = config['file_path']
        ts_token_raw = config['timestamp_column']
        data_columns_raw = config['data_columns']
        sheet_name = config.get('sheet_name', 0)
        headers_map = config.get('headers')  # optional per-file headers mapping

        self.logger.info("Processing file: %s", file_path)

        # Load the Excel file
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Apply optional headers mapping for this file (before column resolution)
        if isinstance(headers_map, dict) and headers_map:
            df = self._apply_headers(df, headers_map)

        # Parse column specs and determine effective skip_rows to preserve alignment
        ts_token, ts_skip = self._parse_col_spec(ts_token_raw)
        parsed_data_cols = []  # list of tuples (label_hint, token, skip)
        max_skip = ts_skip
        for item in data_columns_raw:
            col_token, col_skip = self._parse_col_spec(item)
            # Prefer explicit label in dict; else keep raw string; else token
            if isinstance(item, dict):
                explicit_label = item.get('label') or item.get('as')
                if isinstance(explicit_label, str) and explicit_label.strip() != "":
                    label_hint = explicit_label
                else:
                    label_hint = item if isinstance(item, str) else col_token
            else:
                label_hint = item if isinstance(item, str) else col_token
            parsed_data_cols.append((label_hint, col_token, col_skip))
            if col_skip > max_skip:
                max_skip = col_skip

        # Apply effective skip to entire sheet for consistent row alignment
        if max_skip > 0:
            if any(skip != max_skip for _, _, skip in parsed_data_cols) or ts_skip != max_skip:
                self.logger.warning("Different skip_rows provided; using max skip_rows=%d to keep alignment", max_skip)
            df = df.iloc[max_skip:, :].reset_index(drop=True)

        # Build a filtered DataFrame supporting both named columns and Excel letters (A, B, C, ...)
        available_data_cols = []
        data_dict = {}
        try:
            ts_series = self._get_series_by_spec(df, ts_token)
            data_dict[ts_token] = ts_series
        except KeyError as e:
            raise KeyError(f"Timestamp column not found (by name or letter): {ts_token}") from e

        for label_hint, col_token, _ in parsed_data_cols:
            try:
                ser = self._get_series_by_spec(df, col_token)
                # Determine effective label: if label_hint is just the token/letter, prefer header/series name
                effective_label = label_hint
                # If headers mapping explicitly renames this token, prefer that name
                if isinstance(headers_map, dict) and col_token in headers_map and isinstance(headers_map[col_token], str):
                    mapped = headers_map[col_token].strip()
                    if mapped:
                        effective_label = mapped
                if (effective_label == col_token and self._is_excel_col_letter(col_token)) or (
                    isinstance(effective_label, str) and effective_label not in df.columns and getattr(ser, 'name', None)
                ):
                    effective_label = ser.name

                data_dict[effective_label] = ser
                available_data_cols.append(effective_label)
            except KeyError:
                self.logger.warning("Data column not found (skipping): %s", col_token)

        if not available_data_cols:
            self.logger.warning("No valid data columns found for file: %s. Skipping merge.", file_path)
            return

        df_filtered = pd.DataFrame(data_dict)

        # Convert timestamp column to datetime (supports time-only)
        df_filtered[ts_token] = self._to_datetime_flexible(df_filtered[ts_token])

        # Merge with reference timestamps
        merged_data = self._merge_by_timestamp(
            result_data, df_filtered, ts_token, available_data_cols
        )

        # Update result_data with merged columns
        for col in available_data_cols:
            if col in merged_data.columns:
                result_data[f"{Path(file_path).stem}_{col}"] = merged_data[col]

        self.logger.info("Added %d columns from %s", len(available_data_cols), file_path)

    def _merge_by_timestamp(self, reference_df, data_df, timestamp_col, data_columns):
        """Merge data based on timestamp matching with tolerance."""
        reference_timestamps = reference_df[self.reference_timestamp_column]
        data_timestamps = data_df[timestamp_col]

        # Initialize result columns with NaN
        result_dict = {col: [np.nan] * len(reference_df) for col in data_columns}

        for ref_idx, ref_time in enumerate(reference_timestamps):
            # Find closest timestamp within tolerance
            time_diffs = abs(data_timestamps - ref_time).dt.total_seconds()

            if self.tolerance_seconds > 0:
                # Find timestamps within tolerance
                within_tolerance = time_diffs <= self.tolerance_seconds
                if within_tolerance.any():
                    matching_indices = time_diffs[within_tolerance].index

                    if len(matching_indices) == 1:
                        # Single match - use it directly
                        match_idx = matching_indices[0]
                        for col in data_columns:
                            result_dict[col][ref_idx] = data_df.loc[match_idx, col]
                    else:
                        # Multiple matches - handle duplicates
                        self._handle_duplicate_timestamps(
                            data_df, matching_indices, data_columns, result_dict, ref_idx, ref_time
                        )
            else:
                # Exact match only
                exact_matches = time_diffs == 0
                if exact_matches.any():
                    matching_indices = time_diffs[exact_matches].index

                    if len(matching_indices) == 1:
                        # Single match
                        match_idx = matching_indices[0]
                        for col in data_columns:
                            result_dict[col][ref_idx] = data_df.loc[match_idx, col]
                    else:
                        # Multiple exact matches - handle duplicates
                        self._handle_duplicate_timestamps(
                            data_df, matching_indices, data_columns, result_dict, ref_idx, ref_time
                        )

        return pd.DataFrame(result_dict)

    def _handle_duplicate_timestamps(self, data_df, matching_indices, data_columns,
                                     result_dict, ref_idx, ref_time):
        """Handle multiple rows with the same timestamp."""
        duplicate_data = data_df.loc[matching_indices]

        # Log the duplicate occurrence
        warning_msg = f"Found {len(matching_indices)} duplicate timestamps for {ref_time}"
        self.duplicate_warnings.append(warning_msg)
        self.logger.warning(warning_msg)

        for col in data_columns:
            values = duplicate_data[col].dropna()  # Remove NaN values

            if len(values) == 0:
                # All values are NaN
                result_dict[col][ref_idx] = np.nan
            elif len(values) == 1:
                # Only one non-NaN value
                result_dict[col][ref_idx] = values.iloc[0]
            else:
                # Multiple non-NaN values - apply strategy
                if self.duplicate_strategy == 'mean':
                    # Only calculate mean for numeric data
                    if pd.api.types.is_numeric_dtype(values):
                        result_dict[col][ref_idx] = values.mean()
                    else:
                        result_dict[col][ref_idx] = values.iloc[0]  # First for non-numeric
                elif self.duplicate_strategy == 'first':
                    result_dict[col][ref_idx] = values.iloc[0]
                elif self.duplicate_strategy == 'last':
                    result_dict[col][ref_idx] = values.iloc[-1]
                elif self.duplicate_strategy == 'max':
                    if pd.api.types.is_numeric_dtype(values):
                        result_dict[col][ref_idx] = values.max()
                    else:
                        result_dict[col][ref_idx] = values.iloc[0]  # First for non-numeric
                elif self.duplicate_strategy == 'min':
                    if pd.api.types.is_numeric_dtype(values):
                        result_dict[col][ref_idx] = values.min()
                    else:
                        result_dict[col][ref_idx] = values.iloc[0]  # First for non-numeric
                else:
                    # Default to first
                    result_dict[col][ref_idx] = values.iloc[0]

    def get_summary(self):
        """Get a summary of the synchronization results."""
        if self.synchronized_data is None:
            return "No synchronized data available."

        summary = {
            'total_rows': len(self.synchronized_data),
            'total_columns': len(self.synchronized_data.columns),
            'timestamp_range': {
                'start': self.synchronized_data[self.reference_timestamp_column].min(),
                'end': self.synchronized_data[self.reference_timestamp_column].max()
            },
            'duplicate_strategy_used': self.duplicate_strategy,
            'duplicate_warnings_count': len(self.duplicate_warnings),
            'duplicate_warnings': self.duplicate_warnings[:10],  # Show first 10 warnings
            'missing_data_summary': {}
        }

        # Calculate missing data for each column
        for col in self.synchronized_data.columns:
            if col != self.reference_timestamp_column:
                missing_count = self.synchronized_data[col].isna().sum()
                summary['missing_data_summary'][col] = {
                    'missing_count': missing_count,
                    'missing_percentage': (missing_count / len(self.synchronized_data)) * 100
                }

        return summary


def load_config(config_file):
    """Load configuration from JSON file."""
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config
    except FileNotFoundError:
        print(f"❌ Configuration file '{config_file}' not found.")
        return None
    except json.JSONDecodeError as e:
        print(f"❌ Error parsing JSON configuration: {e}")
        return None
    except (OSError, ValueError) as e:
        print(f"❌ Error loading configuration: {e}")
        return None


def validate_config(config):
    """Validate the configuration structure."""
    required_keys = ['reference_file', 'reference_timestamp_column', 'output_file', 'data_files']

    for key in required_keys:
        if key not in config:
            print(f"❌ Missing required configuration key: '{key}'")
            return False

    # Validate data_files structure
    if not isinstance(config['data_files'], list):
        print("❌ 'data_files' must be a list")
        return False

    for i, file_config in enumerate(config['data_files']):
        required_file_keys = ['file_path', 'timestamp_column', 'data_columns']
        for key in required_file_keys:
            if key not in file_config:
                print(f"❌ Missing required key '{key}' in data_files[{i}]")
                return False

        if not isinstance(file_config['data_columns'], list):
            print(f"❌ 'data_columns' must be a list in data_files[{i}]")
            return False

    return True


def print_summary(summary):
    """Print a formatted summary of the synchronization results."""
    print("\n📊 Synchronization Summary:")
    print(f"   • Total rows: {summary['total_rows']}")
    print(f"   • Total columns: {summary['total_columns']}")
    print(f"   • Time range: {summary['timestamp_range']['start']} to {summary['timestamp_range']['end']}")

    print("\n📈 Data completeness:")
    for col, stats in summary['missing_data_summary'].items():
        completeness = 100 - stats['missing_percentage']
        print(f"   • {col}: {completeness:.1f}% complete ({stats['missing_count']} missing values)")

    # Show duplicate handling information
    if summary['duplicate_warnings_count'] > 0:
        print("\n⚠️  Duplicate timestamps handled:")
        print(f"   • Strategy used: {summary['duplicate_strategy_used']}")
        print(f"   • Total duplicates found: {summary['duplicate_warnings_count']}")
        if summary['duplicate_warnings']:
            print("   • Sample warnings:")
            for warning in summary['duplicate_warnings'][:3]:
                print(f"     - {warning}")
    else:
        print("\n✅ No duplicate timestamps found")


def main():
    """Main function to run the Excel synchronizer."""
    parser = argparse.ArgumentParser(description='Synchronize Excel files based on timestamps')
    parser.add_argument('config', nargs='?', default='config.json',
                        help='Path to JSON configuration file (default: config.json)')

    args = parser.parse_args()

    print("🔄 Excel Timestamp Synchronizer")
    print("=" * 40)

    # Load configuration
    print(f"📖 Loading configuration from: {args.config}")
    config = load_config(args.config)
    if config is None:
        sys.exit(1)

    # Validate configuration
    if not validate_config(config):
        sys.exit(1)

    print("✅ Configuration loaded and validated")

    # Initialize synchronizer
    synchronizer = ExcelTimestampSynchronizer(
        reference_file=config['reference_file'],
        reference_timestamp_column=config['reference_timestamp_column'],
        tolerance_seconds=config.get('tolerance_seconds', 0),
    duplicate_strategy=config.get('duplicate_strategy', 'mean'),
    reference_headers=config.get('reference_headers')
    )

    # Run synchronization
    print("\n🔄 Starting synchronization...")
    print(f"📁 Reference file: {config['reference_file']}")
    print(f"📊 Processing {len(config['data_files'])} data files...")

    success = synchronizer.synchronize_files(config['data_files'], config['output_file'])

    if success:
        print("\n✅ Synchronization completed successfully!")
        print(f"📁 Output saved to: {config['output_file']}")

        # Display summary
        summary = synchronizer.get_summary()
        print_summary(summary)

    else:
        print("\n❌ Synchronization failed. Check the error messages above.")
        sys.exit(1)


if __name__ == "__main__":
    main()
